
# Create your views here.
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.http import HttpResponse, JsonResponse
from .form import SignupForm, LoginForm  # Ensure you import forms correctly
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from .models import Employee, Manager
from .form import SignupForm, LoginForm, EmployeeForm, ManagerForm
from django.urls import reverse
from django.shortcuts import render, redirect, get_object_or_404
from django.core.paginator import Paginator
from django.shortcuts import render, get_object_or_404, redirect
from openpyxl import Workbook

def home(request):
    return render(request, 'home.html')

def user_signup(request):
    if request.method == 'POST':
        form = SignupForm(request.POST)
        if form.is_valid():
            try:
                user = form.save(commit=False)
                user.set_password(form.cleaned_data['password1'])  
                user.save()
                login(request, user)  # Log the user in after successful signup
                return redirect('dashboards')  # Redirect to dashboards
            except Exception as e:
                return HttpResponse(f"Error saving user: {str(e)}")
        else:
            errors = form.errors.as_json()
            return HttpResponse(f"Form is not valid. Please ensure all fields are correctly filled. Errors: {errors}")
    else:
        form = SignupForm()
    return render(request, 'signup.html', {'form': form})



def user_login(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            print(f"Debug: Username - {username}, Password - {password}")  # Debugging line
            user = authenticate(request, username=username, password=password)
            if user is not None:
                if user.is_staff or user.is_superuser:
                    return HttpResponse("Admin users cannot log in from here.")
                login(request, user)
                return redirect('dashboards')  # Redirect to dashboards
            else:
                form.add_error(None, 'Invalid username or password')
    else:
        form = LoginForm()
    return render(request, 'login.html', {'form': form})


def user_logout(request):
    logout(request)
    return redirect('home')

def dashboards(request):
    return render(request, 'dashboards.html')


def Employeee(request):
    # Fetch all employees from the database
    all_employees = Employee.objects.all()

    # Pagination
    paginator = Paginator(all_employees, 10)  # Show 10 employees per page
    page_number = request.GET.get('page')
    employees = paginator.get_page(page_number)

    if request.GET.get('download_excel'):
        # Create a new Excel workbook
        wb = Workbook()

        # Create a new sheet
        ws = wb.active
        ws.title = "Employee Data"

        # Add column headers
        ws.append(['Name', 'Mobile', 'Salary', 'Location', 'Job Role'])

        # Add employee data to the sheet
        for employee in all_employees:
            ws.append([employee.name, employee.mobile, employee.salary, employee.location, employee.job_role])

        # Create the HttpResponse object with the appropriate headers.
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="employee_data.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    return render(request ,'Employee.html', {'employees': employees})

def add_emp(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            # Save the form data to the database
            form.save()
            # Redirect to the Employee page after successful submission
            return HttpResponseRedirect(reverse('Employee'))
    else:
        form = EmployeeForm()
    return render(request ,'add_emp.html', {'form': form})

def add_manager(request):
    if request.method == 'POST':
        form = ManagerForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('Managerr'))
    else:
        form = ManagerForm()
    return render(request, 'add_manager.html', {'form': form})

def Managerr(request):
    all_managers = Manager.objects.all()
    paginator = Paginator(all_managers, 10)
    page_number = request.GET.get('page')
    managers = paginator.get_page(page_number)

    if request.GET.get('download_excel'):
        wb = Workbook()
        ws = wb.active
        ws.title = "Manager Data"
        ws.append(['Name', 'Mobile', 'Salary', 'Section', 'Project Name', 'Department'])
        for manager in all_managers:
            ws.append([manager.name, manager.mobile, manager.salary, manager.section, manager.project_name, manager.department])
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="manager_data.xlsx"'
        wb.save(response)
        return response

    return render(request, 'Managerr.html', {'managers': managers})


def update_emp(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        form = EmployeeForm(request.POST, instance=employee)
        if form.is_valid():
            form.save()
            return redirect('Employee')
    else:
        form = EmployeeForm(instance=employee)
    return render(request, 'update_emp.html', {'form': form})

def delete_emp(request, pk):
    employee = get_object_or_404(Employee, pk=pk)
    if request.method == 'POST':
        employee.delete()
        return redirect('Employee')
    return render(request, 'delete_emp.html', {'employee': employee})


def Managerr(request):
    # Fetch all managers from the database
    all_managers = Manager.objects.all()

    # Pagination
    paginator = Paginator(all_managers, 10)  # Show 10 managers per page
    page_number = request.GET.get('page')
    managers = paginator.get_page(page_number)

    if request.GET.get('download_excel'):
        # Create a new Excel workbook
        wb = Workbook()

        # Create a new sheet
        ws = wb.active
        ws.title = "Manager Data"

        # Add column headers
        ws.append(['Name', 'Mobile', 'Salary', 'Section', 'Project Name', 'Department'])

        # Add manager data to the sheet
        for manager in all_managers:
            ws.append([manager.name, manager.mobile, manager.salary, manager.section, manager.project_name, manager.department])

        # Create the HttpResponse object with the appropriate headers.
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="manager_data.xlsx"'

        # Save the workbook to the response
        wb.save(response)

        return response

    return render(request ,'Managerr.html', {'managers': managers})

def add_manager(request):
    if request.method == 'POST':
        form = ManagerForm(request.POST)
        if form.is_valid():
            form.save()
            return HttpResponseRedirect(reverse('Managerr'))  # Redirect to Managerr page after saving
    else:
        form = ManagerForm()
    return render(request, 'add_manager.html', {'form': form})


def add_emp(request):
    if request.method == 'POST':
        form = EmployeeForm(request.POST)
        if form.is_valid():
            # Save the form data to the database
            form.save()
            # Redirect to the Employee page after successful submission
            return HttpResponseRedirect(reverse('Employee'))
    else:
        form = EmployeeForm()
    return render(request ,'add_emp.html', {'form': form})

def update_man(request, pk):
    manager = get_object_or_404(Manager, pk=pk)
    if request.method == 'POST':
        form = ManagerForm(request.POST, instance=manager)
        if form.is_valid():
            form.save()
            return redirect('Managerr')
    else:
        form = ManagerForm(instance=manager)
    return render(request, 'update_man.html', {'form': form})

def delete_man(request, pk):
    manager = get_object_or_404(Manager, pk=pk)
    if request.method == 'POST':
        manager.delete()
        return redirect('Managerr')
    return render(request, 'delete_man.html', {'manager': manager})